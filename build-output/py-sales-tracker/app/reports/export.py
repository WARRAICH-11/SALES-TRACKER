from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

from sqlalchemy import select, func
from sqlalchemy.orm import Session

from app.data.models import Sale, SaleItem, Product

EXPORTS_DIR = Path(__file__).resolve().parents[2] / "exports"
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def export_sales_to_excel(session: Session, start: date, end: date) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    headers = [
        "Sale ID", "Date", "Product", "Quantity", "Price", "Line Total"
    ]
    ws.append(headers)

    stmt = (
        select(Sale.id, Sale.created_at, Product.name, SaleItem.quantity, SaleItem.price)
        .join(SaleItem, Sale.id == SaleItem.sale_id)
        .join(Product, Product.id == SaleItem.product_id)
        .where(Sale.created_at >= start, Sale.created_at < end)
        .order_by(Sale.created_at.desc())
    )

    for sale_id, created_at, product_name, quantity, price in session.execute(stmt):
        line_total = Decimal(quantity) * Decimal(price)
        ws.append([sale_id, created_at.strftime("%Y-%m-%d %H:%M"), product_name, quantity, float(price), float(line_total)])

    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 20

    out_path = EXPORTS_DIR / f"sales_{start.isoformat()}_{end.isoformat()}.xlsx"
    wb.save(out_path.as_posix())
    return out_path


def export_daily_summary_pdf(session: Session, day: date) -> Path:
    out_path = EXPORTS_DIR / f"daily_summary_{day.isoformat()}.pdf"
    c = canvas.Canvas(out_path.as_posix(), pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 16)
    c.drawString(2 * cm, height - 2 * cm, "Daily Sales Summary")

    c.setFont("Helvetica", 11)
    c.drawString(2 * cm, height - 3 * cm, f"Date: {day.isoformat()}")

    start_dt = day
    end_dt = date.fromordinal(day.toordinal() + 1)

    total_stmt = (
        select(func.sum(SaleItem.quantity * SaleItem.price))
        .join(SaleItem, Sale.id == SaleItem.sale_id)
        .where(Sale.created_at >= start_dt, Sale.created_at < end_dt)
    )
    total = session.execute(total_stmt).scalar() or 0

    c.drawString(2 * cm, height - 4 * cm, f"Total Sales: {float(total):.2f}")

    c.setFont("Helvetica-Bold", 12)
    c.drawString(2 * cm, height - 5.5 * cm, "Top Products")

    top_stmt = (
        select(Product.name, func.sum(SaleItem.quantity))
        .join(SaleItem, Product.id == SaleItem.product_id)
        .join(Sale, Sale.id == SaleItem.sale_id)
        .where(Sale.created_at >= start_dt, Sale.created_at < end_dt)
        .group_by(Product.name)
        .order_by(func.sum(SaleItem.quantity).desc())
        .limit(10)
    )

    y = height - 6.5 * cm
    c.setFont("Helvetica", 11)
    for name, qty in session.execute(top_stmt):
        c.drawString(2 * cm, y, f"{name}: {int(qty)}")
        y -= 0.7 * cm
        if y < 2 * cm:
            c.showPage()
            y = height - 2 * cm

    c.showPage()
    c.save()
    return out_path 